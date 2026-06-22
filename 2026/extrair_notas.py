# -*- coding: utf-8 -*-
"""
Automação de Extração de Valores de Notas Fiscais
Turbo Partners - 2026

Extrai o valor líquido de cada nota fiscal (PDF) organizada por mês e categoria,
gerando uma planilha Excel com resumos por pasta e por mês.
"""

import os
import re
import sys
from pathlib import Path

import pandas as pd
import pdfplumber
from openpyxl.styles import Font, PatternFill, Alignment, numbers
from openpyxl.utils import get_column_letter

# ============================================================================
# CONFIGURAÇÃO
# ============================================================================

BASE_DIR = Path(r"g:/Drives compartilhados/BackOffice Turbo Partners/01 - Contabilidade/FISCAL/TURBO/2026")
OUTPUT_FILE = BASE_DIR / "relatorio_notas_2026.xlsx"
CAMBIO_USD_BRL = 6.00  # Câmbio fixo USD -> BRL (ajustar conforme necessário)

# Extensões suportadas
PDF_EXTENSIONS = {".pdf"}
IMAGE_EXTENSIONS = {".jfif", ".jpg", ".jpeg", ".png", ".bmp", ".tiff"}
SKIP_FILES = {"desktop.ini", "extrair_notas.py", "relatorio_notas_2026.xlsx"}

# ============================================================================
# FUNÇÕES DE PARSING DE VALORES
# ============================================================================

def parse_brl(value_str: str) -> float | None:
    """Converte string de valor BRL (1.234,56) para float."""
    if not value_str:
        return None
    cleaned = value_str.strip().replace(" ", "")
    # Formato BRL: pontos separam milhares, vírgula separa decimais
    # Ex: 1.234,56 ou 234,56 ou 1.234.567,89
    if "," in cleaned:
        cleaned = cleaned.replace(".", "").replace(",", ".")
    else:
        # Sem vírgula - pode ser inteiro ou formato com ponto decimal
        # Se tem ponto e os últimos dígitos após o ponto são <= 2, é decimal
        if "." in cleaned:
            parts = cleaned.split(".")
            if len(parts[-1]) <= 2:
                # Pode ser decimal (ex: 140.00) - manter como está
                pass
            else:
                # Milhares sem decimal (ex: 1.234) - remover pontos
                cleaned = cleaned.replace(".", "")
    try:
        val = float(cleaned)
        if val < 0:
            return None
        return val
    except (ValueError, TypeError):
        return None


def parse_usd(value_str: str) -> float | None:
    """Converte string de valor USD (1,234.56) para float."""
    if not value_str:
        return None
    cleaned = value_str.strip().replace(" ", "").replace(",", "").replace("$", "")
    try:
        val = float(cleaned)
        if val < 0:
            return None
        return val
    except (ValueError, TypeError):
        return None


# ============================================================================
# PADRÕES REGEX PARA EXTRAÇÃO DE VALORES
# ============================================================================

# Cada padrão é uma tupla: (regex_compilado, moeda, descrição)
# A ordem importa - padrões mais específicos primeiro

PATTERNS_BRL = [
    # --- GRUPO 1: Valor Líquido (prioridade máxima) ---

    # 1a. "Valor Líquido da NFS-e" com R$ na mesma linha
    (re.compile(
        r'Valor\s+L[ií]quido\s+da\s+NFS-?e\s*[:=]?\s*R\$\s*([\d.,]+)',
        re.IGNORECASE
    ), "BRL", "Valor Liquido da NFS-e (mesma linha)"),

    # 1b. "Valor Líquido da NFS-e" com R$ na próxima linha (Vila Velha/Cariacica)
    (re.compile(
        r'Valor\s+L[ií]quido\s+da\s+NFS-?e\s*[-\s]*R\$([\d.,]+)',
        re.IGNORECASE | re.DOTALL
    ), "BRL", "Valor Liquido da NFS-e (prox linha)"),

    # 1c. "VALOR LÍQUIDO DA NOTA FISCAL : R$" (Lauro de Freitas)
    (re.compile(
        r'VALOR\s+L[IÍ]QUIDO\s+DA\s+NOTA\s+FISCAL\s*[:=]?\s*R\$\s*([\d.,]+)',
        re.IGNORECASE
    ), "BRL", "Valor Liquido da Nota Fiscal"),

    # 1d. "VALOR LÍQUIDO: R$" (L.A.X e similares)
    (re.compile(
        r'VALOR\s+L[IÍ]QUIDO\s*[:=]\s*R\$\s*([\d.,]+)',
        re.IGNORECASE
    ), "BRL", "VALOR LIQUIDO"),

    # 1e. "Valor líquido = R$" (PATRIMONIUM)
    (re.compile(
        r'Valor\s+l[ií]quido\s*=\s*R\$\s*([\d.,]+)',
        re.IGNORECASE
    ), "BRL", "Valor liquido = R$"),

    # 1f. "Vl. Líquido da Nota Fiscal R$" (Ribeirão Preto)
    (re.compile(
        r'Vl\.\s*L[ií]quido\s+(?:da\s+)?Nota\s+Fiscal\s*R?\$?\s*([\d.,]+)',
        re.IGNORECASE
    ), "BRL", "Vl. Liquido da Nota Fiscal"),

    # 1g. "Valor Líquido" seguido de número na mesma linha (GISS - Paulínia, Santos)
    (re.compile(
        r'Valor\s+L[ií]quido\s+([\d][\d.,]*)',
        re.IGNORECASE
    ), "BRL", "Valor Liquido (sem R$)"),

    # --- GRUPO 2: Valor Total da Nota / NFS-e ---

    # 2a. VALOR TOTAL RECEBIDO (São Paulo - Caju)
    (re.compile(
        r'VALOR\s+TOTAL\s+RECEBIDO\s*=?\s*R\$\s*([\d.,]+)',
        re.IGNORECASE
    ), "BRL", "Valor Total Recebido"),

    # 2b. "VALOR TOTAL DA NOTA FISCAL : R$" (Lauro de Freitas)
    (re.compile(
        r'VALOR\s+TOTAL\s+DA\s+NOTA\s+FISCAL\s*[:=]?\s*R\$\s*([\d.,]+)',
        re.IGNORECASE
    ), "BRL", "Valor Total da Nota Fiscal"),

    # 2c. "Valor Total dos Serviços" seguido de valor na próxima linha (NFS-e Avulsa)
    (re.compile(
        r'Valor\s+Total\s+dos\s+Servi[çc]os\s*\n\s*([\d.,]+)',
        re.IGNORECASE
    ), "BRL", "Valor Total dos Servicos (prox linha)"),

    # 2d. Valor Total da Nota com R$ na mesma linha (Osasco, etc.)
    (re.compile(
        r'Valor\s+Total\s+da\s+Nota\s*[:=]?\s*R\$\s*([\d.,]+)',
        re.IGNORECASE
    ), "BRL", "Valor Total da Nota (com R$)"),

    # 2e. Valor Total da Nota seguido de número na mesma linha (sem R$)
    (re.compile(
        r'Valor\s+Total\s+da\s+Nota\s*[:=]\s*([\d][\d.,]*)',
        re.IGNORECASE
    ), "BRL", "Valor Total da Nota (sem R$)"),

    # 2f. "Valor Total da NFS-e X.XXX,XX" (sem R$, formato inline)
    (re.compile(
        r'Valor\s+Total\s+da\s+NFS-?e\s+([\d.,]+)',
        re.IGNORECASE
    ), "BRL", "Valor Total da NFS-e (inline)"),

    # 2g. Valor Total do Serviço (NFS-e várias prefeituras)
    (re.compile(
        r'VALOR\s+TOTAL\s+DO\s+SERVI[CÇ]O\s*[:=]?\s*R\$\s*([\d.,]+)',
        re.IGNORECASE
    ), "BRL", "Valor Total do Servico"),

    # 2f. Valor Total do Documento (DARF - impostos com valor na mesma linha)
    (re.compile(
        r'Valor\s+Total\s+do\s+Documento\s*[:=]?\s*R?\$?\s*([\d.,]+)',
        re.IGNORECASE
    ), "BRL", "Valor Total do Documento"),

    # --- GRUPO 3: Formatos concatenados (Vitória) ---

    # 3a. NFS-e Vitória concatenado - VALORTOTALDANFS-E seguido de R$
    (re.compile(
        r'VALORTOTALDANFS-?E.*?R\$([\d.,]+)',
        re.IGNORECASE | re.DOTALL
    ), "BRL", "VALORTOTALDANFS-E (Vitoria)"),

    # 3b. ValordoServiço concatenado (Vitória)
    (re.compile(
        r'Valordoservi[çc]o.*?R\$([\d.,]+)',
        re.IGNORECASE | re.DOTALL
    ), "BRL", "ValordoServico (Vitoria)"),

    # --- GRUPO 4: NFS-e Vila Velha / Cariacica (com espaços) ---

    # 4a. "VALOR TOTAL DA NFS-E" (com espaços) seguido de R$ na próxima linha
    (re.compile(
        r'VALOR\s+TOTAL\s+DA\s+NFS-?E.*?R\$([\d.,]+)',
        re.IGNORECASE | re.DOTALL
    ), "BRL", "VALOR TOTAL DA NFS-E (espacos)"),

    # 4b. "Valor do Serviço" com espaços + R$ (Vila Velha fallback)
    (re.compile(
        r'Valor\s+do\s+Servi[çc]o\s+Desconto\s+Condicionado.*?R\$([\d.,]+)',
        re.IGNORECASE | re.DOTALL
    ), "BRL", "Valor do Servico (Vila Velha)"),

    # --- GRUPO 5: Faturas e contas ---

    # 5a. Valor fatura (Café Express)
    (re.compile(
        r'Valor\s+fatura\s*[:=]?\s*R\$\s*([\d.,]+)',
        re.IGNORECASE
    ), "BRL", "Valor fatura"),

    # 5b. Valor devido (ADMOOH)
    (re.compile(
        r'Valor\s+devido\s*[:=]?\s*R\$\s*([\d.,]+)',
        re.IGNORECASE
    ), "BRL", "Valor devido"),

    # 5c. Total a Pagar (Vivo, concessionárias)
    (re.compile(
        r'Total\s+a\s+Pagar\s*[-:=]?\s*R?\$?\s*([\d.,]+)',
        re.IGNORECASE
    ), "BRL", "Total a Pagar"),

    # 5d. Total a Recolher (ISS municipal)
    (re.compile(
        r'Total\s+a\s+Recolher\s*[:=]?\s*R\$\s*([\d.,]+)',
        re.IGNORECASE
    ), "BRL", "Total a Recolher"),

    # 5e. TOTAL A PAGAR (variação)
    (re.compile(
        r'TOTAL\s+A\s+PAGAR\s*([\d.,]+)',
        re.IGNORECASE
    ), "BRL", "TOTAL A PAGAR"),

    # 5f. Valor Total: (Facebook, genérico)
    (re.compile(
        r'Valor\s+Total\s*[:=]\s*R?\$?\s*([\d.,]+)',
        re.IGNORECASE
    ), "BRL", "Valor Total"),

    # 5g. "Valor: X.XXX,XX" (DAS/DARF no rodapé do boleto)
    (re.compile(
        r'\nValor:\s*([\d.,]+)',
        re.IGNORECASE
    ), "BRL", "Valor: (boleto)"),

    # --- GRUPO 6: EDP energia ---

    # 6a. EDP - formato "MES/ANO DD/MM/AAAA R$ X.XXX,XX"
    (re.compile(
        r'(?:JAN|FEV|MAR|ABR|MAI|JUN|JUL|AGO|SET|OUT|NOV|DEZ)/\d{4}\s+\d{2}/\d{2}/\d{4}\s+R\$\s*([\d.,]+)',
        re.IGNORECASE
    ), "BRL", "EDP Energia"),

    # --- GRUPO 7: NFe de produto (Cervejaria etc.) ---
    # Nota: para NFe, usamos função especial em extract_value_from_text

    # --- GRUPO 8: Boleto Imobiliária (valor standalone) ---

    # 8a. "R$ XX.XXX,XX" precedido de "1" (quantidade no boleto)
    (re.compile(
        r'\n1\s+R\$\s*([\d.,]+)',
        re.IGNORECASE
    ), "BRL", "Boleto (1 R$)"),
]

PATTERNS_USD = [
    # 9a. Amount due $XX.XX (Invoices internacionais)
    (re.compile(
        r'Amount\s+due\s*\$?\s*([\d.,]+)',
        re.IGNORECASE
    ), "USD", "Amount due (USD)"),

    # 9b. Total amount XXX.XX USD
    (re.compile(
        r'Total\s+amount\s*([\d.,]+)\s*USD',
        re.IGNORECASE
    ), "USD", "Total amount USD"),
]


def _smart_parse_value(raw_value: str) -> float | None:
    """Parse valor detectando automaticamente formato BRL ou USD."""
    if not raw_value:
        return None
    # Para valores com vírgula E ponto, detectar formato
    if "," in raw_value and "." in raw_value:
        comma_pos = raw_value.rfind(",")
        dot_pos = raw_value.rfind(".")
        if comma_pos < dot_pos:
            # Formato 100,818.86 (vírgula = milhares, ponto = decimal)
            return parse_usd(raw_value)
        else:
            # Formato 1.234,56 (ponto = milhares, vírgula = decimal)
            return parse_brl(raw_value)
    return parse_brl(raw_value)


def extract_value_from_text(text: str) -> tuple[float | None, str, str]:
    """
    Tenta extrair o valor da nota a partir do texto do PDF.
    Retorna (valor_em_brl, moeda_original, padrão_usado).
    """
    if not text:
        return None, "", ""

    # Tentar padrões BRL primeiro
    for pattern, currency, desc in PATTERNS_BRL:
        match = pattern.search(text)
        if match:
            raw_value = match.group(1)
            value = _smart_parse_value(raw_value)

            if value is not None and value > 0:
                return value, "BRL", desc

    # Tentar padrões USD
    for pattern, currency, desc in PATTERNS_USD:
        match = pattern.search(text)
        if match:
            raw_value = match.group(1)
            value = parse_usd(raw_value)
            if value is not None and value > 0:
                brl_value = value * CAMBIO_USD_BRL
                return brl_value, "USD", desc

    # --- Fallbacks especiais para formatos complexos ---

    # Bom Despacho: "Valor Líquido" em header de tabela, valor na linha de dados
    # Formato: "PIS Outras Retenções ... Valor Líquido\n0,00 0,00 ... 1.550,00"
    vl_header = re.search(
        r'Valor\s+L[ií]quido\s*\n([\d.,\s]+)',
        text, re.IGNORECASE
    )
    if vl_header:
        data_line = vl_header.group(1).strip()
        numbers = re.findall(r'[\d.,]+', data_line)
        if numbers:
            value = _smart_parse_value(numbers[-1])  # último número = Valor Líquido
            if value is not None and value > 0:
                return value, "BRL", "Valor Liquido (tabela, ultimo)"

    # NFe de produto: "VALOR TOTAL DA NOTA" seguido de linha com múltiplos R$
    # O último R$ é o valor total
    nfe_match = re.search(
        r'VALOR\s+TOTAL\s+DA\s+NOTA\s*\n(.*)',
        text, re.IGNORECASE
    )
    if nfe_match:
        line = nfe_match.group(1)
        all_values = re.findall(r'R\$\s*([\d.,]+)', line)
        if all_values:
            value = _smart_parse_value(all_values[-1])  # último valor
            if value is not None and value > 0:
                return value, "BRL", "VALOR TOTAL DA NOTA (NFe ultimo)"

    # NFS-e Avulsa: "Valor Total da Nota" como header + última coluna na próx linha
    avulsa_match = re.search(
        r'Valor\s+Total\s+da\s+Nota\s*\n([\d.,\s]+)',
        text, re.IGNORECASE
    )
    if avulsa_match:
        data_line = avulsa_match.group(1).strip()
        numbers = re.findall(r'[\d.,]+', data_line)
        if numbers:
            value = _smart_parse_value(numbers[-1])
            if value is not None and value > 0:
                return value, "BRL", "Valor Total da Nota (tabela, ultimo)"

    # DAS: "Valor Total do Documento" em linha separada, valor na próxima
    das_match = re.search(
        r'Valor\s+Total\s+do\s+Documento\s*\n\s*([\d.,]+)',
        text, re.IGNORECASE
    )
    if das_match:
        value = _smart_parse_value(das_match.group(1))
        if value is not None and value > 0:
            return value, "BRL", "Valor Total do Documento (prox linha)"

    # Boleto Imobiliária: "R$ XX.XXX,XX" (exclui valores de multa/juros/tributos)
    # Filtra linhas que contêm "tributo", "multa", "juros", "aproximado"
    all_rs = []
    for line in text.split("\n"):
        ll = line.lower()
        if any(k in ll for k in ["tributo", "multa", "juros", "aproximado", "vencimento"]):
            continue
        for m in re.finditer(r'R\$\s*([\d.]+,\d{2})', line):
            all_rs.append(m.group(1))
    if all_rs:
        best_val = 0.0
        for raw in all_rs:
            v = _smart_parse_value(raw)
            if v and v > best_val:
                best_val = v
        if best_val > 0:
            return best_val, "BRL", "Maior R$ encontrado (fallback)"

    return None, "", ""


# ============================================================================
# FUNÇÕES DE EXTRAÇÃO DE PDF
# ============================================================================

def extract_text_from_pdf(filepath: Path) -> tuple[str, str]:
    """
    Extrai texto de um PDF.
    Retorna (texto, status).
    Status pode ser: "OK", "PROTEGIDO", "SEM_TEXTO", "ERRO".
    """
    try:
        with pdfplumber.open(str(filepath)) as pdf:
            all_text = []
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    all_text.append(text)
            if all_text:
                return "\n".join(all_text), "OK"
            else:
                return "", "SEM_TEXTO"
    except Exception as e:
        error_msg = str(e).lower()
        if "password" in error_msg or "encrypt" in error_msg:
            return "", "PROTEGIDO"
        return "", f"ERRO: {e}"


# ============================================================================
# SCAN E PROCESSAMENTO
# ============================================================================

def scan_directory(base_dir: Path) -> list[dict]:
    """
    Percorre a estrutura de pastas e extrai valores de cada nota.
    Estrutura esperada: base_dir / MES / CATEGORIA / arquivo.pdf
    """
    results = []

    # Listar pastas de meses (ordenadas)
    month_dirs = sorted([
        d for d in base_dir.iterdir()
        if d.is_dir() and re.match(r'\d{2}\s*-\s*', d.name)
    ])

    total_files = 0
    processed = 0

    # Contar total de arquivos primeiro
    for month_dir in month_dirs:
        for category_dir in sorted(month_dir.iterdir()):
            if not category_dir.is_dir():
                continue
            for f in category_dir.iterdir():
                if f.name.lower() in SKIP_FILES:
                    continue
                if f.suffix.lower() in PDF_EXTENSIONS | IMAGE_EXTENSIONS:
                    total_files += 1

    print(f"Total de arquivos encontrados: {total_files}")
    print("-" * 60)

    for month_dir in month_dirs:
        month_name = month_dir.name  # Ex: "01 - JANEIRO"

        for category_dir in sorted(month_dir.iterdir()):
            if not category_dir.is_dir():
                continue

            category_name = category_dir.name  # Ex: "Fixo", "Freelancer"

            for filepath in sorted(category_dir.iterdir()):
                if filepath.name.lower() in SKIP_FILES:
                    continue

                ext = filepath.suffix.lower()

                # Imagens - não suportado
                if ext in IMAGE_EXTENSIONS:
                    results.append({
                        "mes": month_name,
                        "pasta": category_name,
                        "arquivo": filepath.name,
                        "valor_brl": None,
                        "moeda_original": "",
                        "padrao_usado": "",
                        "status": "IMAGEM - NÃO SUPORTADO",
                    })
                    processed += 1
                    print(f"  [{processed}/{total_files}] {filepath.name} -> IMAGEM")
                    continue

                if ext not in PDF_EXTENSIONS:
                    continue

                # Extrair texto do PDF
                text, pdf_status = extract_text_from_pdf(filepath)

                if pdf_status == "PROTEGIDO":
                    results.append({
                        "mes": month_name,
                        "pasta": category_name,
                        "arquivo": filepath.name,
                        "valor_brl": None,
                        "moeda_original": "",
                        "padrao_usado": "",
                        "status": "PDF PROTEGIDO",
                    })
                    processed += 1
                    print(f"  [{processed}/{total_files}] {filepath.name} -> PROTEGIDO")
                    continue

                if pdf_status == "SEM_TEXTO":
                    results.append({
                        "mes": month_name,
                        "pasta": category_name,
                        "arquivo": filepath.name,
                        "valor_brl": None,
                        "moeda_original": "",
                        "padrao_usado": "",
                        "status": "SEM TEXTO (IMAGEM)",
                    })
                    processed += 1
                    print(f"  [{processed}/{total_files}] {filepath.name} -> SEM TEXTO")
                    continue

                if pdf_status.startswith("ERRO"):
                    results.append({
                        "mes": month_name,
                        "pasta": category_name,
                        "arquivo": filepath.name,
                        "valor_brl": None,
                        "moeda_original": "",
                        "padrao_usado": "",
                        "status": pdf_status,
                    })
                    processed += 1
                    print(f"  [{processed}/{total_files}] {filepath.name} -> {pdf_status}")
                    continue

                # Extrair valor do texto
                valor, moeda, padrao = extract_value_from_text(text)

                if valor is not None:
                    results.append({
                        "mes": month_name,
                        "pasta": category_name,
                        "arquivo": filepath.name,
                        "valor_brl": round(valor, 2),
                        "moeda_original": moeda,
                        "padrao_usado": padrao,
                        "status": "OK",
                    })
                    processed += 1
                    moeda_tag = f" ({moeda})" if moeda == "USD" else ""
                    print(f"  [{processed}/{total_files}] {filepath.name} -> R$ {valor:,.2f}{moeda_tag}")
                else:
                    results.append({
                        "mes": month_name,
                        "pasta": category_name,
                        "arquivo": filepath.name,
                        "valor_brl": None,
                        "moeda_original": "",
                        "padrao_usado": "",
                        "status": "VALOR NÃO ENCONTRADO",
                    })
                    processed += 1
                    print(f"  [{processed}/{total_files}] {filepath.name} -> NÃO ENCONTRADO")

    return results


# ============================================================================
# GERAÇÃO DO EXCEL
# ============================================================================

def generate_excel(results: list[dict], output_path: Path):
    """Gera planilha Excel com abas de detalhado, resumo por pasta, resumo mensal e erros."""

    df = pd.DataFrame(results)

    # --- Aba Detalhado ---
    df_detail = df[["mes", "pasta", "arquivo", "valor_brl", "moeda_original", "status"]].copy()
    df_detail.columns = ["Mês", "Pasta", "Arquivo", "Valor (R$)", "Moeda Original", "Status"]

    # --- Aba Resumo por Pasta ---
    df_ok = df[df["status"] == "OK"]
    df_err = df[df["status"] != "OK"]

    summary_pasta = []
    for (mes, pasta), group in df.groupby(["mes", "pasta"], sort=True):
        ok_count = (group["status"] == "OK").sum()
        err_count = (group["status"] != "OK").sum()
        total = group.loc[group["status"] == "OK", "valor_brl"].sum()
        summary_pasta.append({
            "Mês": mes,
            "Pasta": pasta,
            "Qtd Notas": ok_count,
            "Total (R$)": round(total, 2),
            "Erros": err_count,
        })
    df_pasta = pd.DataFrame(summary_pasta)

    # --- Aba Resumo Mensal ---
    summary_month = []
    for mes, group in df.groupby("mes", sort=True):
        ok_count = (group["status"] == "OK").sum()
        err_count = (group["status"] != "OK").sum()
        total = group.loc[group["status"] == "OK", "valor_brl"].sum()
        summary_month.append({
            "Mês": mes,
            "Qtd Notas": ok_count,
            "Total (R$)": round(total, 2),
            "Notas com Erro": err_count,
        })
    # Adicionar linha de TOTAL
    total_ok = sum(r["Qtd Notas"] for r in summary_month)
    total_val = sum(r["Total (R$)"] for r in summary_month)
    total_err = sum(r["Notas com Erro"] for r in summary_month)
    summary_month.append({
        "Mês": "TOTAL",
        "Qtd Notas": total_ok,
        "Total (R$)": round(total_val, 2),
        "Notas com Erro": total_err,
    })
    df_month = pd.DataFrame(summary_month)

    # --- Aba Erros ---
    df_errors = df[df["status"] != "OK"][["mes", "pasta", "arquivo", "status"]].copy()
    df_errors.columns = ["Mês", "Pasta", "Arquivo", "Motivo"]

    # --- Escrever Excel ---
    with pd.ExcelWriter(str(output_path), engine="openpyxl") as writer:
        df_detail.to_excel(writer, sheet_name="Detalhado", index=False)
        df_pasta.to_excel(writer, sheet_name="Resumo por Pasta", index=False)
        df_month.to_excel(writer, sheet_name="Resumo Mensal", index=False)
        df_errors.to_excel(writer, sheet_name="Erros", index=False)

        # Formatação
        wb = writer.book

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        money_format = '#,##0.00'
        total_font = Font(bold=True, size=12)
        total_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Formatar cabeçalho
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            # Auto-ajustar largura das colunas
            for col_idx, col in enumerate(ws.columns, 1):
                max_length = 0
                col_letter = get_column_letter(col_idx)
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = min(max_length + 4, 60)

            # Formatar colunas de valor como moeda
            for col_idx, cell in enumerate(ws[1], 1):
                if cell.value and "R$" in str(cell.value):
                    for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                        for c in row:
                            if c.value is not None:
                                c.number_format = money_format

        # Formatar linha de TOTAL no Resumo Mensal
        ws_month = wb["Resumo Mensal"]
        last_row = ws_month.max_row
        for cell in ws_month[last_row]:
            cell.font = total_font
            cell.fill = total_fill

    print(f"\nPlanilha gerada: {output_path}")


# ============================================================================
# MAIN
# ============================================================================

def main():
    print("=" * 60)
    print("EXTRAÇÃO DE VALORES - NOTAS FISCAIS 2026")
    print(f"Câmbio USD/BRL: R$ {CAMBIO_USD_BRL:.2f}")
    print("=" * 60)
    print()

    results = scan_directory(BASE_DIR)

    print()
    print("=" * 60)
    print("GERANDO PLANILHA EXCEL...")
    print("=" * 60)

    generate_excel(results, OUTPUT_FILE)

    # Resumo no terminal
    df = pd.DataFrame(results)
    ok_count = (df["status"] == "OK").sum()
    err_count = (df["status"] != "OK").sum()
    total_val = df.loc[df["status"] == "OK", "valor_brl"].sum()

    print()
    print("=" * 60)
    print("RESUMO")
    print("=" * 60)
    print(f"  Notas processadas com sucesso: {ok_count}")
    print(f"  Notas com erro/não encontrado: {err_count}")
    print(f"  Valor total extraído: R$ {total_val:,.2f}")
    print()

    # Mostrar erros
    errors = df[df["status"] != "OK"]
    if not errors.empty:
        print("NOTAS COM PROBLEMAS:")
        for _, row in errors.iterrows():
            print(f"  [{row['status']}] {row['pasta']}/{row['arquivo']}")


if __name__ == "__main__":
    main()
